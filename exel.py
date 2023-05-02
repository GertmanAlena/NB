import openpyxl

import logger as log
from openpyxl.comments import Comment
import search_name_file_month as snfm
import datetime as DT

filePath = 'D:/studies/BotNotaryMy/Notarius/Zapis.xlsx'
def x_file(day, notarius):
    """
    метод выбора свободного времени для записи
    :param d: выбранный день
    :param notarius: выбранный нотариус
    :return: free_time список свободного времени
    """
    try:
        sheet = activ_list(notarius, day) # получаем нужный акривный лист нужного нотариуса
        max_rows = sheet.max_row
        max_cols = sheet.max_column
        free_time = data_z(sheet, day, notarius, max_rows, max_cols) # выбираем свободное время
        return free_time
    except Exception as e:
        log.log_exel_x_file(e, "x_file")
        return False
def data_z(sheet, day, notarius, max_rows, max_cols):
    """
    метод поиска нужной ячейки для записи принимает параметры
    :param sheet: активный лист
    :param max_rows: максимальное количество строк в листе
    :param max_cols: максимальное количество колонок в листе
    :param d: день для записи
    :param time_work: время работы
    :return:  free_time список со свободным временем для записи
    """
    free_time = []
    time_work = ''
    try:
        for dat in range(1, max_rows+1):
            z = str(sheet.cell(row=dat, column=1).value)
            date_table = z.split(' ')[0]
            if sheet.cell(row=dat, column=1).value is not None and day == date_table:
                if z.split(' ')[1] == 'воскресенье':
                    free_time.append('воскресенье')
                    return free_time
                else:
                    row_day = dat
                    for i in range(1, max_cols + 1):
                        if sheet.cell(row=1, column=i).value == notarius:
                            column_not = i
                            time_work = sheet.cell(row=row_day, column=column_not).value
                            free_time.append(time_work)
                            if z.split(' ')[1] == 'понедельник' or z.split(' ')[1] == 'суббота':
                                for k in range(row_day + 1, row_day + 6):
                                    if sheet.cell(row=k, column=i).value is None:
                                        free_time.append(sheet.cell(row=k, column=1).value.strftime("%H:%M"))
                            else:
                                for k in range(row_day+1, row_day + 12):
                                    if sheet.cell(row=k, column=i).value is None:
                                        free_time.append(sheet.cell(row=k, column=1).value.strftime("%H:%M"))

        if free_time[0] == 'выходной':
            free_time.append('выходной')
            return free_time
        else:
            work_end = time_work.split("-")[1]
        if day == DT.datetime.now().strftime("%d.%m.%Y") and DT.datetime.now().strftime("%H:%M") > work_end:
            print(DT.datetime.now().strftime("%H:%M"), " > ", work_end)
            free_time.append('завершён')
            return free_time
        else:
            return free_time
    except Exception as e:
        print("error", e)
def activ_list(notarius, day):
    month = day.split('.')[1]
    try:
        workbook = openpyxl.load_workbook(filePath)
        worksheet = workbook[snfm.search_month(month)]
        return worksheet
    except Exception as e:
        print("e activ_list", e)
        log.activ_list(e, notarius, day)
        return False

def save_file(time, notarius, day, power_of_attorney, name, last_name, tel):
    wb = openpyxl.load_workbook(filePath)
    mes = snfm.search_month(day.split(".")[1])
    worksheet = wb[mes]
    """
    метод записи и сохранения
    :param wb: диерктория нотариуса
    :param worksheet: активный лист месяца для записи
    :param notarius: нотариус
    :param time: время для записи
    :param day: день для записи
    :return:
    """

    res = str(name) + " " + str(last_name) + " " + str(tel)
    max_rows = worksheet.max_row
    max_cols = worksheet.max_column
    try:
        for dat in range(1, max_rows+1):
            # ищем нужную дату в строке
            z = str(worksheet.cell(row=dat, column=1).value)
            x = z.split(" ")[0]
            if worksheet.cell(row=dat, column=1).value is not None and\
                    day == x:
                row_day = dat
        for i in range(1, max_cols + 1):
            if worksheet.cell(row=1, column=i).value == notarius:
                column_not = i
        for k in range(row_day + 1, row_day + 12):
            if worksheet.cell(row=k, column=1).value is not None and time == worksheet.cell(row=k,
                                                                                            column=1).value.strftime(
                        "%H:%M"):
                worksheet.cell(row=k, column=column_not).value = power_of_attorney
                comment = Comment(text=res, author='-')
                worksheet.cell(row=k, column=column_not).comment = comment
                wb.save(filePath)
                return True
    except Exception as e:
        log.save_file(e, "save_file")
        print(e)
        return False
def search(telephone):
    """
    метод поиска записи
    :param telephone: принимает номер телефона гражданина по которому будет искать куда он записан
    :return: дату время и к какому нотариусу записан гражданин
    """
    result_all = []
    result = ''
    try:
        wb = openpyxl.load_workbook(filePath)
        day = DT.datetime.now().strftime("%d.%m.%Y")
        start_month = DT.datetime.now().strftime("%m")

        for l in range(int(start_month), 13):          # проходим по листам в файле
            mes = snfm.search_month(str(l))
            worksheet = wb[mes]
            max_rows = worksheet.max_row
            max_cols = worksheet.max_column
            for i in range(3, max_rows + 1):            # по строкам
                for j in range(1, max_cols + 1):            # по столбцам
                    result_zapis = []
                    if worksheet.cell(row=i, column=j).comment:         # если комментарий есть

                        com = worksheet.cell(row=i, column=j).comment
                        result = str(com).split(" ")
                        tel = ""
                        for k in range(0, len(result)):
                            if result[k].isnumeric():
                                tel = result[k]
                                if str(telephone) == tel:
                                    notar = worksheet.cell(row=1, column=j).value
                                    action = str(worksheet.cell(row=i, column=j).value)
                                    time_zapis = str(worksheet.cell(row=i, column=1).value)
                                    if time_zapis == "08:00:00":
                                        a = str(worksheet.cell(row=i-1, column=1).value.split(" ")[0])
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "09:00:00":
                                        a = str(worksheet.cell(row=i-2, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "10:00:00":
                                        a = str(worksheet.cell(row=i-3, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "11:00:00":
                                        a = str(worksheet.cell(row=i-4, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "12:00:00":
                                        a = str(worksheet.cell(row=i-5, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "13:00:00":
                                        a = str(worksheet.cell(row=i-6, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "14:00:00":
                                        a = str(worksheet.cell(row=i-7, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "15:00:00":
                                        a = str(worksheet.cell(row=i-8, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "16:00:00":
                                        a = str(worksheet.cell(row=i-9, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "17:00:00":
                                        a = str(worksheet.cell(row=i-10, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    elif time_zapis == "18:00:00":
                                        a = str(worksheet.cell(row=i-11, column=1).value)
                                        data_zapis = str(a.split(" ")[0])
                                    if data_zapis.split(".")[2] == \
                                            DT.datetime.now().strftime("%d.%m.%Y").split(".")[2] and \
                                            data_zapis.split(".")[1] == \
                                            DT.datetime.now().strftime("%d.%m.%Y").split(".")[
                                                1] and \
                                            data_zapis.split(".")[0] >= \
                                            DT.datetime.now().strftime("%d.%m.%Y").split(".")[
                                                0]:
                                        result_zapis.append(action)
                                        result_zapis.append(notar)
                                        result_zapis.append(time_zapis)
                                        result_zapis.append(data_zapis)
                                        print('result_zapis ', result_zapis)
                                        result_all.append(result_zapis)
                                    elif data_zapis.split(".")[2] == \
                                            DT.datetime.now().strftime("%d.%m.%Y").split(".")[2] and \
                                            data_zapis.split(".")[1] > \
                                            DT.datetime.now().strftime("%d.%m.%Y").split(".")[
                                                1]:
                                        result_zapis.append(action)
                                        result_zapis.append(notar)
                                        result_zapis.append(time_zapis)
                                        result_zapis.append(data_zapis)
                                        print('result_zapis ', result_zapis)
                                        result_all.append(result_zapis)
                            else:
                                continue
                            # сравниваем день месяц и год, чтобы не попались прошедшие даты

        print('result_all ', result_all)
        return result_all

    except Exception as e:
        log.search(e)
        print(e)
        return False

def delete_file(mess):

    day = mess.split(' ')[7]
    time = mess.split(' ')[5]
    notarius = mess.split(' ')[1]
    wb = openpyxl.load_workbook(filePath)
    mes = snfm.search_month(day.split(".")[1])
    worksheet = wb[mes]
    """
    метод удаления записи и сохранения
    :param wb: файл xslx
    :param worksheet: активный лист месяца для записи
    :param notarius: нотариус
    :param time: время для записи
    :param day: день для записи
    :return:
    """

    max_rows = worksheet.max_row
    max_cols = worksheet.max_column
    try:
        for dat in range(1, max_rows+1):
            # ищем нужную дату в строке
            z = str(worksheet.cell(row=dat, column=1).value)
            x = z.split(" ")[0]
            if worksheet.cell(row=dat, column=1).value is not None and\
                    day == x:
                row_day = dat
        # ищем нотариуса
        for i in range(1, max_cols + 1):
            if worksheet.cell(row=1, column=i).value == notarius:
                column_not = i
        # ищем нужное время
        for k in range(row_day + 1, row_day + 12):
            x=worksheet.cell(row=k, column=1).value.strftime("%H:%M:%S")
            if worksheet.cell(row=k, column=1).value is not None and time == x:
                worksheet.cell(row=k, column=column_not).value = None
                # comment = Comment(text=None, author=None)
                worksheet.cell(row=k, column=column_not).comment = None
                wb.save(filePath)
                return True
    except Exception as e:
        log.save_file(e, "save_file")
        print(e)
        return False